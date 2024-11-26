from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, numbers, colors
import datetime
import sqlite3
from hard_coding import *



def createExcel(db, last, first):
    
    db_name = 'databases/' + str(db) + '.db'
    print(db_name)
    conn = sqlite3.connect(db_name)
    cur = conn.cursor()
    cur.execute('''CREATE TABLE IF NOT EXISTS plants (name TEXT, qty TEXT, size TEXT, cost TEXT, plant_type TEXT)''')
    cur.execute('''CREATE TABLE IF NOT EXISTS services (name TEXT, material TEXT, mat_cost TEXT, manhours TEXT)''')
    cur = cur.execute('''SELECT * FROM plants''')
    data = cur.fetchall()
    # for i in data:
    #     print(i)

    createWorkbook(db_name)  
    conn.close()



def createWorkbook(db):

    wb = Workbook()

    ws = wb.active
    ws.column_dimensions['A'].width= 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    ws['A1'] = datetime.date.today()

    thick_border = Side(border_style="thick", color="000000")
    thin_border = Side(border_style="thin", color="000000")

    #PLANT ROW HEADERS

    ws['A7'] = 'Notes:'
    ws['B7']= 'qty'
    ws['B7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['B7'].alignment = Alignment(horizontal='center')
    ws['B7'].font = Font(bold=True, size= 12)
    ws['C7'] = 'descriptions'
    ws['C7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['C7'].alignment = Alignment(horizontal='center')
    ws['C7'].font = Font(bold=True, size= 12)
    ws['D7'] = 'unit'
    ws['D7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['D7'].alignment = Alignment(horizontal='center')
    ws['D7'].font = Font(bold=True, size= 12)
    ws['E7'] = 'unit cost'
    ws['E7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['E7'].alignment = Alignment(horizontal='center')
    ws['E7'].font = Font(bold=True, size= 12)
    ws['F7'] = 'Ext. Plant Cost'
    ws['F7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['F7'].alignment = Alignment(horizontal='center')
    ws['F7'].font = Font(bold=True, size= 12)
    ws['G7'] = 'labor factor'
    ws['G7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['G7'].alignment = Alignment(horizontal='center')
    ws['G7'].font = Font(bold=True, size= 12)
    ws['H7'] = 'man hours'
    ws['H7'].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['H7'].alignment = Alignment(horizontal='center')
    ws['H7'].font = Font(bold=True, size= 12)

    conn = sqlite3.connect(db)

    cur = conn.cursor()
    

    #PLANT LABOR FACTOR DATA

    cur.execute('''CREATE TABLE IF NOT EXISTS labor_factors (con_qrt TEXT, con_gal TEXT, con_2gal TEXT, con_3gal TEXT, con_5gal TEXT, con_7gal TEXT, con_10gal TEXT, con_15gal TEXT, con_25gal TEXT,
                    dec_15 TEXT, dec_20 TEXT, dec_25 TEXT, dec_30 TEXT, dec_35 TEXT, dec_40 TEXT, dec_45 TEXT, dec_50 TEXT, dec_60 TEXT, dec_70 TEXT,
                    ever_4 TEXT, ever_5 TEXT, ever_6 TEXT, ever_7 TEXT, ever_8 TEXT, ever_10 TEXT, ever_12 TEXT, ever_14 TEXT,
                    sh_12 TEXT, sh_15 TEXT, sh_18 TEXT, sh_24 TEXT, sh_30 TEXT, sh_36 TEXT, sh_48 TEXT
                    )''') 
    laborfactor_data = cur.execute('''SELECT * FROM labor_factors ORDER BY ROWID DESC LIMIT 1''').fetchone() 
    if laborfactor_data == None:
        # print('try')
        cur.execute('''INSERT INTO labor_factors VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ''',(base_factors_dict['quart'], base_factors_dict['1gal'],base_factors_dict['2gal'], base_factors_dict['3gal'], base_factors_dict['5gal'], base_factors_dict['7gal'], base_factors_dict['10gal'], base_factors_dict['15gal'], base_factors_dict['25gal'],
                         base_factors_dict['one5inch'], base_factors_dict['twoinch'], base_factors_dict['two5inch'], base_factors_dict['threeinch'], base_factors_dict['three5inch'], base_factors_dict['fourinch'], base_factors_dict['four5inch'], base_factors_dict['fiveinch'], base_factors_dict['sixinch'], base_factors_dict['seveninch'],
                         base_factors_dict['four5'], base_factors_dict['five6'], base_factors_dict['six7'], base_factors_dict['seven8'],base_factors_dict['eight10'], base_factors_dict['ten12'], base_factors_dict['twelve14'], base_factors_dict['fourteen16'],
                         base_factors_dict['twelve'], base_factors_dict['fifteen'], base_factors_dict['eighteen'], base_factors_dict['twentyfour'], base_factors_dict['thirty'],base_factors_dict['thirtysix'], base_factors_dict['fortyeight']))
        conn.commit()
        laborfactor_data = cur.execute('''SELECT * FROM labor_factors ORDER BY ROWID DESC LIMIT 1''').fetchone()
        # print('except')
    # print('labor factor data is', laborfactor_data)
    
    db_labor_factors = {
                "quart" : laborfactor_data[0],
                "1gal" : laborfactor_data[1],
                "2gal" : laborfactor_data[2],
                "3gal" : laborfactor_data[3],
                "5gal"  : laborfactor_data[4],
                "7gal" : laborfactor_data[5],
                "10gal" : laborfactor_data[6],
                "15gal" : laborfactor_data[7],
                "25gal" : laborfactor_data[8],
                "dec_15" : laborfactor_data[9],
                "dec_20" : laborfactor_data[10],
                "dec_25" : laborfactor_data[11],
                "dec_30" : laborfactor_data[12],
                "dec_35" : laborfactor_data[13],
                "dec_40" : laborfactor_data[14],
                "dec_45" : laborfactor_data[15],
                "dec_50" : laborfactor_data[16],
                "dec_60" : laborfactor_data[17],
                "dec_70" : laborfactor_data[18],
                "ev_4" : laborfactor_data[19],
                "ev_5" : laborfactor_data[20],
                "ev_6" : laborfactor_data[21],
                "ev_7" : laborfactor_data[22],
                "ev_8" : laborfactor_data[23],
                "ev_10" : laborfactor_data[24],
                "ev_12" : laborfactor_data[25],
                "ev_14" : laborfactor_data[26],
                "sh_12" : laborfactor_data[27],
                "sh_15" : laborfactor_data[28],
                "sh_18" : laborfactor_data[29],
                "sh_24" : laborfactor_data[30],
                "sh_30" : laborfactor_data[31],
                "sh_36" : laborfactor_data[32],
                "sh_48" : laborfactor_data[33]
                }
    
    #SERVICE LABOR FACTOR DATA
# 
    cur.execute('''CREATE TABLE IF NOT EXISTS service_labor_factors (mulch TEXT, soil TEXT, stone TEXT, flagstone TEXT, sixbysixbyeight_footer TEXT, sixbysixbyeight_course TEXT, paver TEXT, ads_4inchpipe TEXT,
                    tilling TEXT, sod_prepped TEXT, sod_unprepped TEXT, sod_prepped_1wide TEXT, sod_prepped_3wide TEXT, sodcutter TEXT,
                    six_upright TEXT, eight_upright TEXT, guywire_2ft TEXT, turnbuckle TEXT
                    )''')
    service_laborfactor_data = cur.execute('''SELECT * FROM service_labor_factors ORDER BY ROWID DESC LIMIT 1''').fetchone() 
    if service_laborfactor_data == None:
        # print('try')
        cur.execute('''INSERT INTO service_labor_factors VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                     ''',(base_service_factors["mulch_1yard"], base_service_factors["soil_1yard"],base_service_factors["stone_1yard"],base_service_factors["flagstone_100sqft_4inchbase"],base_service_factors["sixbysixbyeight_footer"],base_service_factors["sixbysixbyeight_course"],base_service_factors["paver_100sqft_4inchbase"],base_service_factors["pipe_4inchx10ft"],
                        base_service_factors["tilling_100sqft"],base_service_factors["sod_500sqft_preppped"],base_service_factors["sod_500sqft_unprepped"],base_service_factors["sod_prepped_1wide"],base_service_factors["sod_prepped_3wide"],base_service_factors["sodcutter_100sqft"],
                          base_service_factors["six_upright"],base_service_factors["eight_upright"],base_service_factors["guywire_2ft"],base_service_factors["turnbuckle"],  ))
        conn.commit()
        service_laborfactor_data = cur.execute('''SELECT * FROM service_labor_factors ORDER BY ROWID DESC LIMIT 1''').fetchone()
        # print('except')
    # print(' service     labor factor data is', service_laborfactor_data)
    

    #ROWS OF PLANTS
    # print(db_labor_factors)

    plant_data = cur.execute('''SELECT * FROM plants''').fetchall()

    # if no plants then print NO PLANTS
    plantrows = len(plant_data)
    if plantrows == 0:
        plantrows = 1
        ws['B8'] = 'NO PLANTS'

    for i in plant_data:
        this_row = str(plant_data.index(i) + 8)
        qty_col = 'B' + this_row
        desc_col = 'C' +this_row
        unit_col = 'D' +this_row
        unit_cost_col = 'E' +this_row
        ext_cost_col = 'F' +this_row
        labor_factor_col = 'G' +this_row
        manhour_col = 'H' +this_row
        ws[qty_col] = float(i[1])
        ws[qty_col].alignment = Alignment(horizontal='center')
        ws[qty_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[desc_col] = i[0]
        ws[desc_col].alignment = Alignment(horizontal='center')
        ws[desc_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[unit_col] = i[2]
        ws[unit_col].alignment = Alignment(horizontal='center')
        ws[unit_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[unit_cost_col] = float(i[3])
        ws[unit_cost_col].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws[unit_cost_col].alignment = Alignment(horizontal='center')
        ws[unit_cost_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[ext_cost_col] = float(i[1]) * float(i[3])
        ws[ext_cost_col].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws[ext_cost_col].alignment = Alignment(horizontal='center')
        ws[ext_cost_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[labor_factor_col] = 0
        ws[labor_factor_col].alignment = Alignment(horizontal='center')
        ws[labor_factor_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[manhour_col] = 0
        ws[manhour_col].alignment = Alignment(horizontal='center')
        ws[manhour_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)



    # Row of SERVICE HEADERS
        
    service_header_row = plantrows + 8
    
    
    ws['B' + str(service_header_row)]= 'Service Name'
    ws['B' + str(service_header_row)].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['B' + str(service_header_row)].alignment = Alignment(horizontal='center')
    ws['B' + str(service_header_row)].font = Font(bold=True, size= 12)
    ws['C' + str(service_header_row)] = 'Materials'
    ws['C' + str(service_header_row)].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['C' + str(service_header_row)].alignment = Alignment(horizontal='center')
    ws['C' + str(service_header_row)].font = Font(bold=True, size= 12)
    ws['D' + str(service_header_row)] = 'Material Cost'
    ws['D' + str(service_header_row)].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['D' + str(service_header_row)].alignment = Alignment(horizontal='center')
    ws['D' + str(service_header_row)].font = Font(bold=True, size= 12)
    ws['E' + str(service_header_row)] = 'Change'
    ws['E' + str(service_header_row)].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['E' + str(service_header_row)].alignment = Alignment(horizontal='center')
    ws['E' + str(service_header_row)].font = Font(bold=True, size= 12)
    ws['F' + str(service_header_row)] = 'Extended Mat Cost'
    ws['F' + str(service_header_row)].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['F' + str(service_header_row)].alignment = Alignment(horizontal='center')
    ws['F' + str(service_header_row)].font = Font(bold=True, size= 12)
    ws['G' + str(service_header_row)] = 'labor factor'
    ws['G' + str(service_header_row)].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['G' + str(service_header_row)].alignment = Alignment(horizontal='center')
    ws['G' + str(service_header_row)].font = Font(bold=True, size= 12)
    ws['H' + str(service_header_row)] = 'man hours'
    ws['H' + str(service_header_row)].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws['H' + str(service_header_row)].alignment = Alignment(horizontal='center')
    ws['H' + str(service_header_row)].font = Font(bold=True, size= 12)



    #ROWS OF SERVICES
    service_data = cur.execute('''SELECT * FROM services''').fetchall()

    service_rows = len(service_data)

    if service_rows == 0:
        service_rows = 1
        
        ws['B' + str(plantrows + 9)] = 'NO SERVICES'

    for i in service_data:
        this_row = str(service_data.index(i) + plantrows + 9)
        name_col = 'B' + this_row
        mat_col = 'C' +this_row
        mat_cost_col = 'D' +this_row
        unit_cost_col = 'E' +this_row
        ext_cost_col = 'F' +this_row
        labor_factor_col = 'G' +this_row
        manhour_col = 'H' +this_row
        ws[name_col] = i[0]
        ws[name_col].alignment = Alignment(horizontal='center')
        ws[name_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[mat_col] = i[1]
        ws[mat_col].alignment = Alignment(horizontal='center')
        ws[mat_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[unit_cost_col] = i[1]
        ws[unit_cost_col].alignment = Alignment(horizontal='center')
        ws[unit_cost_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[mat_cost_col] = float(i[2])
        ws[mat_cost_col].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws[mat_cost_col].alignment = Alignment(horizontal='center')
        ws[mat_cost_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[ext_cost_col] = float(i[2]) * 2
        ws[ext_cost_col].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        ws[ext_cost_col].alignment = Alignment(horizontal='center')
        ws[ext_cost_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[labor_factor_col] = 0
        ws[labor_factor_col].alignment = Alignment(horizontal='center')
        ws[labor_factor_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)
        ws[manhour_col] = i[3]
        ws[manhour_col].alignment = Alignment(horizontal='center')
        ws[manhour_col].border = Border(top=thin_border, left=thin_border, right=thin_border, bottom=thin_border)





    # ROWS OF TOTALS
    direct_row = 'C' + str(plantrows+ service_rows + 10)
    ws[direct_row] = 'DIRECT COST LABOR'
    ws[direct_row].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thin_border)
    ws[direct_row].font = Font(bold=False, size= 9)
    ws[direct_row].alignment = Alignment(horizontal='center')
    direct_mat_row = 'C' + str(plantrows+ service_rows + 11)
    ws[direct_mat_row] = 'DIRECT COST MATERIALS(Materials, Tax, Freight)'
    ws[direct_mat_row].border = Border(top=thin_border, left=thick_border, right=thick_border, bottom=thin_border)
    ws[direct_mat_row].font = Font(bold=False, size= 8)
    ws[direct_mat_row].alignment = Alignment(horizontal='center')
    billable_eqip_row = 'C' + str(plantrows + service_rows + 12)
    ws[billable_eqip_row] = 'Billable Equipment Rate'
    ws[billable_eqip_row].border = Border(top=thin_border, left=thick_border, right=thick_border, bottom=thin_border)
    ws[billable_eqip_row].font = Font(bold=False, size= 9)
    ws[billable_eqip_row].alignment = Alignment(horizontal='center')
    total_direct_row = 'C' + str(plantrows + service_rows + 13)
    ws[total_direct_row] = 'TOTAL DIRECT COST'
    ws[total_direct_row].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws[total_direct_row].font = Font(bold=False, size= 9)
    ws[total_direct_row].alignment = Alignment(horizontal='center')
    desired_markup_row = 'C' + str(plantrows + service_rows + 15)
    ws[desired_markup_row] = 'Enter Desired Mat Markup %'
    ws[desired_markup_row].border = Border(top=thick_border, left=thick_border, right=thick_border, bottom=thick_border)
    ws[desired_markup_row].font = Font(bold=True, size= 9)
    ws[desired_markup_row].alignment = Alignment(horizontal='center')
    ws[desired_markup_row].fill = PatternFill('solid', start_color="ffff00")



    wb.save("workbook.xlsx")

